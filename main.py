# author: bingcheng li
# date: 2021/11/25 5:52 下午
# email: 740194688@qq.com
# coding=utf-8

import openpyxl, os, shutil, time
from decimal import Decimal
from openpyxl.styles import Alignment, Font


def create_sheet():
    wb2 = openpyxl.load_workbook(file_path)
    new_head = []
    new_data = []
    first_sheet_name = wb2.sheetnames[0]
    ws = wb2[first_sheet_name]
    titles = ws[1]

    for t in titles:
        new_head.append(t.value)
        try:
            val = Decimal(obj[t.column_letter]).quantize(Decimal('0.00'))
            new_data.append(val)
        except KeyError:
            new_data.append('')

    new_ws = wb2.create_sheet('汇总表')
    new_ws.append(new_head)
    new_ws.append(new_data)
    rows = new_ws.rows
    for row in rows:
        for grid in row:
            grid.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    # print(f'new_data:{new_data}')

    try:
        wb2.save(file_path)
        wb2.close()
        print('success: 去查看文件并保存')
        return False
    except PermissionError as e:
        print(f'{e} | 文件为打开状态，关闭文件后重来一次')
        return True
    except Exception as e:
        print(f'{e} | 保存文件异常')
        return True


def get_data():
    for sheet in all_sheet:
        if sheet[:3] != '汇总表':
            cur_sheet = wb[sheet]
            for kk in keys:
                col = cur_sheet[kk]
                for cell in col:
                    if cell.row != 1:
                        try:
                            value = float(cell.value)
                            obj[kk] += Decimal(value)
                        except (ValueError, TypeError) as e:
                            if cell.value is not None:
                                log(f'sheetName: {sheet} | pos: {cell.coordinate}  | err: {e} | value: {cell.value}')
    else:
        # print(obj)
        # wb.save('copy.xlsx')
        wb.close()
        # os.unlink(copy_file)
        return create_sheet()


def log(msg):
    print(f'msg: {msg}')
    with open('./log.txt', 'a') as log_file:
        log_file.write(msg + '\n')


def show_files():
    cur_dir_files = []
    p = os.getcwd()
    dirs = os.listdir(p)

    for file in dirs:
        if os.path.isfile(p + '\\' + file):
            if file.split(sep='.')[1] == 'xlsx':
                cur_dir_files.append(file)
    for f in cur_dir_files:
        print(f'{cur_dir_files.index(f)}：{f}')
    print('----------------------------------------------------')
    while True:
        i = input('\n选择强牛牛要统计的文件编号：')
        try:
            name = cur_dir_files[int(i)]
            file_exist = os.path.exists('./' + name)
            if file_exist:
                break
            else:
                print('file is not exist')
        except Exception as e:
            print(f'输入错误: {e}')
    return name


while True:
    print('----------------------------------------------------')
    print('[================== 强牛牛的自动统计EXCLE')
    print('[================== 把文件放在当前目录')
    print('[================== 数据无价谨慎操作,建议使用复制品进行\n')

    file_path = show_files()
    try:
        # shutil.copy(file_path, copy_file)
        wb = openpyxl.load_workbook('./' + file_path, data_only=True)
    except Exception as e:
        log(f'file path|type error：{e}')
        continue
    else:
        val = input('输入强牛牛要计算的所有列：')
        if val.isdigit():
            print('error: 乖乖要计算的列只能输字母。。。')
            continue

        keys = list(val.upper())
        confirm = input(f'确认列：{keys}  y/n ? ')
        if confirm.upper() == 'Y':
            obj = {}
            for k in keys:
                obj[k] = 0
            all_sheet = wb.sheetnames
            result = get_data()
            if not result:
                time.sleep(10)
                break
        else:
            continue
